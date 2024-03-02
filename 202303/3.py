from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import requests

class Vac(APIView):
    def get(self, request, *args, **kwargs):
        # Fetch data from Class1 API
        response_class1 = requests.get('url/to/class1/api/', params=request.GET)
        data_class1 = response_class1.json()
        
        # Fetch data from Class2 API
        response_class2 = requests.get('url/to/class2/api/', params=request.GET)
        data_class2 = response_class2.json()
        
        # Load Excel template
        wb = load_workbook('path/to/template.xlsx')
        
        # Write data from Class1 to first table
        ws1 = wb['Sheet1']  # Update with actual sheet name
        start_row_ws1 = 2  # Assuming data starts from row 2 (below headers)
        for row in dataframe_to_rows(pd.DataFrame(data_class1), index=False, header=False):
            for idx, value in enumerate(row, start=start_row_ws1):
                ws1.cell(row=idx, column=1).value = value
            start_row_ws1 += 1
        
        # Write data from Class2 to second table
        ws2 = wb['Sheet2']  # Update with actual sheet name
        start_row_ws2 = 2  # Assuming data starts from row 2 (below headers)
        for row in dataframe_to_rows(pd.DataFrame(data_class2), index=False, header=False):
            for idx, value in enumerate(row, start=start_row_ws2):
                ws2.cell(row=idx, column=1).value = value
            start_row_ws2 += 1
        
        # Save modified Excel file
        modified_file_path = 'path/to/modified_template.xlsx'
        wb.save(modified_file_path)
        
        # Serve modified Excel file as a download
        with open(modified_file_path, 'rb') as file:
            response = Response(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="modified_template.xlsx"'
            return response