# Load Excel file
excel_file = 'your_excel_file.xlsx'
excel_writer = pd.ExcelWriter(excel_file, engine='openpyxl')
excel_writer.book = load_workbook(excel_file)
excel_writer.sheets = dict((ws.title, ws) for ws in excel_writer.book.worksheets)

# Write DataFrame to Excel, append to existing table
df_json.to_excel(excel_writer, sheet_name='Sheet1', index=False, startrow=1, header=False)

# Save changes
excel_writer.save()




import pandas as pd

# Read JSON file into pandas DataFrame
df_json = pd.read_json('your_json_file.json')

# Load Excel file and select the sheet
excel_file = 'your_excel_file.xlsx'
sheet_name = 'Sheet1'  # Replace 'Sheet1' with the actual sheet name
excel_data = pd.read_excel(excel_file, sheet_name=sheet_name)

# Update the existing table in Excel with the new data
with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl') as writer:
    writer.book = openpyxl.load_workbook(excel_file)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    df_json.to_excel(writer, sheet_name=sheet_name, startrow=len(excel_data)+1, index=False, header=False)
    
    
    
    
    
    
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Load the existing Excel file
excel_file = "your_excel_file.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

# Sample DataFrame with new data
new_data = {
    'Name': ['Tom', 'Emma'],
    'Age': [40, 28],
    'City': ['Seattle', 'San Francisco']
}
df_new = pd.DataFrame(new_data)

# Determine the range of the existing table (assuming it starts from A1)
table_range = f"A1:{chr(ord('A') + len(df_new.columns) - 1)}1"

# Write new data from DataFrame to worksheet
for r in dataframe_to_rows(df_new, index=False, header=False):
    ws.append(r)

# Update the table range to include the new data
new_table_range = f"{table_range.split(':')[0]}:{chr(ord('A') + len(df_new.columns) - 1)}{len(df_new) + 1}"

# Save the workbook
wb.save(excel_file)




from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Load the existing Excel file
excel_file = "your_excel_file.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

# Sample DataFrame with new data
new_data = {
    'Name': ['Tom', 'Emma'],
    'Age': [40, 28],
    'City': ['Seattle', 'San Francisco']
}
df_new = pd.DataFrame(new_data)

# Get the existing table range and name
table_name = "MyTable"  # Replace with the actual table name
existing_table = ws.tables[table_name]
existing_table_range = existing_table.ref

# Write new data from DataFrame to worksheet below the header
for r in dataframe_to_rows(df_new, index=False, header=False):
    ws.append(r)

# Update the table range to include the new data
existing_table_range_start, existing_table_range_end = existing_table_range.split(':')
new_table_range_end = f"{existing_table_range_end[0]}{int(existing_table_range_end[1:]) + len(df_new)}"
new_table_range = f"{existing_table_range_start}:{new_table_range_end}"

# Update the existing table's range
existing_table.ref = new_table_range

# Save the workbook
wb.save(excel_file)

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import re

# Load the existing Excel file
excel_file = "your_excel_file.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

# Sample DataFrame with new data
new_data = {
    'Name': ['Tom', 'Emma'],
    'Age': [40, 28],
    'City': ['Seattle', 'San Francisco']
}
df_new = pd.DataFrame(new_data)

# Get the existing table range and name
table_name = "MyTable"  # Replace with the actual table name
existing_table = ws.tables[table_name]
existing_table_range = existing_table.ref

# Write new data from DataFrame to worksheet below the header
for r in dataframe_to_rows(df_new, index=False, header=False):
    ws.append(r)

# Update the table range to include the new data
table_range_start, table_range_end = existing_table_range.split(':')

# Extract column letter(s) and row number using regular expressions
start_col_letter = re.findall(r'[A-Z]+', table_range_start)[0]
end_col_letter = re.findall(r'[A-Z]+', table_range_end)[0]
start_row_number = int(re.findall(r'\d+', table_range_start)[0])
end_row_number = int(re.findall(r'\d+', table_range_end)[0])

# Determine new end column letter (handle cases where column index exceeds 26)
new_end_col_letter = chr(ord(end_col_letter[-1]) + len(df_new.columns))
if len(new_end_col_letter) > 1:
    new_end_col_letter = chr(ord(end_col_letter[-2]) + 1) + chr(ord(end_col_letter[-1]) + len(df_new.columns) % 26)

# Construct the new table range
new_table_range = f"{start_col_letter}{start_row_number}:{new_end_col_letter}{end_row_number + len(df_new)}"

# Update the existing table's range
existing_table.ref = new_table_range

# Save the workbook
wb.save(excel_file)






from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
import pandas as pd

# Load the existing Excel file
excel_file = "your_excel_file.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

# Sample DataFrame with new data
new_data = {
    'Name': ['Tom', 'Emma'],
    'Age': [40, 28],
    'City': ['Seattle', 'San Francisco']
}
df_new = pd.DataFrame(new_data)

# Get the existing table range and name
table_name = "MyTable"  # Replace with the actual table name
existing_table = ws.tables[table_name]
existing_table_range = existing_table.ref

# Convert the start and end coordinates of the table range to row and column indices
start_row, start_col = coordinate_to_tuple(existing_table_range.split(":")[0])
end_row, end_col = coordinate_to_tuple(existing_table_range.split(":")[1])

# Write new data from DataFrame to worksheet below the header
for r in dataframe_to_rows(df_new, index=False, header=False):
    ws.append(r)

# Update the table range to include the new data
new_end_col = end_col + len(df_new.columns)
new_end_row = end_row + len(df_new)
new_table_range = f"{coordinate_to_column_string(start_col)}{start_row}:{coordinate_to_column_string(new_end_col)}{new_end_row}"

# Update the existing table's range
existing_table.ref = new_table_range

# Save the workbook
wb.save(excel_file)


