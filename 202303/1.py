from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from rest_framework import generics
from .models import YourModel
from .serializers import YourModelSerializer

class Vac(generics.ListCreateAPIView):
    queryset = YourModel.objects.all()
    serializer_class = YourModelSerializer
    
    def get(self, request, *args, **kwargs):
        filter_parameter = request.GET.get('filter_parameter')
        
        # Filter queryset based on filter_parameter
        if filter_parameter:
            self.queryset = self.queryset.filter(parameter=filter_parameter)
        
        # Serialize queryset data
        serializer = self.serializer_class(self.queryset, many=True)
        data = serializer.data
        
        # Create a DataFrame from serialized data
        df = pd.DataFrame(data)
        
        # Load Excel template
        wb = load_workbook('path/to/your/template.xlsx')
        ws = wb.active
        
        # Populate Excel template with data
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        
        # Configure PivotTable
        table = Table(displayName="Data", ref="A1:D1")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Set PivotTable data source
        pivot_table = ws.pivot_tables.create('PivotTable', ws['A1'], ws['E1'], location='PivotTable')
        pivot_table_data = pivot_table.table.ref
        pivot_table_data_source = f'{ws.title}!{pivot_table_data}'
        pivot_table.data_source = pivot_table_data_source
        
        # Set PivotTable refresh settings
        pivot_table.refresh_on_load = True
        
        # Save modified Excel file
        modified_file_path = 'path/to/modified_template.xlsx'
        wb.save(modified_file_path)
        
        # Serve modified Excel file as a download
        with open(modified_file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="modified_template.xlsx"'
            return response